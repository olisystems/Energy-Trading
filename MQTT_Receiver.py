import paho.mqtt.client as mqtt
import  time
import openpyxl
import os.path
broker="m23.cloudmqtt.com"
port=10094
username=""
password=""


if (os.path.exists("smartpi_data.xlsx") == False):
	wb=openpyxl.Workbook()
	ws=wb.active
	ws['A1']="OLI Systems GmbH"
	wb.save('smartpi_data.xlsx')
wb = openpyxl.load_workbook('smartpi_data.xlsx')
ws = wb.active
def on_connect(client, userdata, flags, rc):
	if rc==0:
		print("Connected with result code"+str(rc))
		client.subscribe("OLI_Power/House/smartpi12345")
		client.subscribe("OLI_Energy/House/smartpi12345")
	else:
		print("Bad connection code=",rc)

def on_message(client, userdata, msg):
	print(msg.topic+" "+str(msg.payload))
	ws.append([str(msg.topic),str(msg.payload)])
	wb.save("smartpi_data.xlsx")


client = mqtt.Client("sub")
client.username_pw_set(username, password)
client.on_connect = on_connect
client.on_message = on_message

print("Connecting to broker", broker)
client.connect(broker,port)
time.sleep(1)
print("in main loop")
client.loop_forever()
client.disconnect()
